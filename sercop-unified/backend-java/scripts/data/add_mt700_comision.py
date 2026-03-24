#!/usr/bin/env python3
import openpyxl
from datetime import datetime

# Cargar el archivo Excel
file_path = '/Users/admin/Documents/capacitacion/claude/claudecode/globalcmx/backend/src/main/resources/comisiones-swift-config.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print('Agregando configuración de comisiones para MT700 EMISSION_LC_IMPORT...')

# Encontrar la última fila con datos
last_row = ws.max_row

# Agregar nuevas reglas para MT700 EMISSION_LC_IMPORT
# Estructura: tipoMensaje, evento, montoMin, montoMax, moneda, paisOrigen, paisDestino, comisionFija, comisionPorcentaje, comisionMinima, comisionMaxima

nuevas_reglas = [
    # MT700 EMISSION_LC_IMPORT - USD - Monto pequeño
    ['MT700', 'EMISSION_LC_IMPORT', 0, 50000, 'USD', 'MX', 'MX', 0, 0.125, 100, 500],
    # MT700 EMISSION_LC_IMPORT - USD - Monto grande
    ['MT700', 'EMISSION_LC_IMPORT', 50000, 999999999, 'USD', 'MX', 'MX', 0, 0.1, 500, 5000],
    # MT700 EMISSION_LC_IMPORT - EUR - Monto pequeño
    ['MT700', 'EMISSION_LC_IMPORT', 0, 50000, 'EUR', 'MX', 'MX', 0, 0.125, 120, 600],
    # MT700 EMISSION_LC_IMPORT - EUR - Monto grande
    ['MT700', 'EMISSION_LC_IMPORT', 50000, 999999999, 'EUR', 'MX', 'MX', 0, 0.1, 600, 6000],
    # MT700 EMISSION_LC_IMPORT - MXN - Monto pequeño
    ['MT700', 'EMISSION_LC_IMPORT', 0, 1000000, 'MXN', 'MX', 'MX', 0, 0.125, 2000, 10000],
    # MT700 EMISSION_LC_IMPORT - MXN - Monto grande
    ['MT700', 'EMISSION_LC_IMPORT', 1000000, 999999999, 'MXN', 'MX', 'MX', 0, 0.1, 10000, 100000],
]

# Agregar cada regla
for i, regla in enumerate(nuevas_reglas, 1):
    fila = last_row + i
    for col, valor in enumerate(regla, 1):
        ws.cell(row=fila, column=col, value=valor)
    print(f'  Agregada regla {i}: {regla[0]} - {regla[1]} - {regla[4]} ({regla[2]:,} - {regla[3]:,})')

# Guardar el archivo
wb.save(file_path)
print(f'\n✅ Archivo guardado exitosamente: {file_path}')
print(f'   Total de reglas agregadas: {len(nuevas_reglas)}')
print(f'   Nueva última fila: {last_row + len(nuevas_reglas)}')

# Crear backup
backup_path = file_path.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
wb.save(backup_path)
print(f'   Backup creado: {backup_path}')
