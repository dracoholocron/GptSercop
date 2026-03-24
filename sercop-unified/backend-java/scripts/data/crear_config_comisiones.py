#!/usr/bin/env python3
"""
Script para generar archivo Excel de configuración de comisiones SWIFT
Compatible con Drools Decision Tables
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def crear_archivo_comisiones():
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comisiones SWIFT"

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Metadatos de Drools (filas 1-8)
    metadata = [
        ["RuleSet", "com.globalcmx.comisiones"],
        ["Import", "com.globalcmx.api.dto.swift.MensajeSWIFT"],
        ["Import", "com.globalcmx.api.dto.comision.ConfiguracionComision"],
        ["Sequential", "true"],
        ["", ""],
        ["RuleTable", "Configuración de Comisiones SWIFT"],
    ]

    for i, row in enumerate(metadata, start=1):
        for j, value in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            if i <= 6 and j == 1:
                cell.font = Font(bold=True, size=10)

    # Fila 7: Headers de la tabla de decisión
    headers = [
        "CONDITION",
        "CONDITION",
        "CONDITION",
        "CONDITION",
        "CONDITION",
        "CONDITION",
        "CONDITION",
        "ACTION",
        "ACTION",
        "ACTION",
        "ACTION"
    ]

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Fila 8: Tipos de objeto y bindings
    # Para CONDITIONS: repetir el tipo de objeto en cada columna
    # Para ACTIONS: declarar el binding con $ una vez, luego celdas vacías
    object_types = [
        "MensajeSWIFT",                # Para tipoMensaje
        "MensajeSWIFT",                # Para evento
        "MensajeSWIFT",                # Para monto mínimo
        "MensajeSWIFT",                # Para monto máximo
        "MensajeSWIFT",                # Para moneda
        "MensajeSWIFT",                # Para paisOrigen
        "MensajeSWIFT",                # Para paisDestino
        "$c : ConfiguracionComision",  # Declaración del binding con $ para ACTIONs
        "",                            # Empty - mismo objeto
        "",                            # Empty - mismo objeto
        ""                             # Empty - mismo objeto
    ]

    for col_num, obj_type in enumerate(object_types, start=1):
        cell = ws.cell(row=8, column=col_num, value=obj_type)
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Fila 9: Constraints (snippets)
    # Para CONDITIONS: usar solo el nombre del campo (sin el prefijo del objeto)
    # Para ACTIONS: usar el binding $c.método con cast a Double
    constraints = [
        "tipoMensaje == \"$param\"",
        "evento == \"$param\"",
        "monto >= $param",
        "monto < $param",
        "moneda == \"$param\"",
        "paisOrigen == \"$param\"",
        "paisDestino == \"$param\"",
        "$c.setComisionFija(Double.valueOf($param))",
        "$c.setComisionPorcentaje(Double.valueOf($param))",
        "$c.setComisionMinima(Double.valueOf($param))",
        "$c.setComisionMaxima(Double.valueOf($param))"
    ]

    for col_num, constraint in enumerate(constraints, start=1):
        cell = ws.cell(row=9, column=col_num, value=constraint)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = border
        cell.font = Font(size=9, italic=True)

    # Datos de ejemplo - Mensajes SWIFT de Comercio Exterior
    data_examples = [
        # === COBRANZAS DOCUMENTARIAS (MT4xx) ===
        # MT400 - Aviso de Cobranza
        ["MT400", "EMISION", "0", "50000", "USD", "MX", "US", "150", "0", "100", "500"],
        ["MT400", "EMISION", "50000", "999999999", "USD", "MX", "US", "0", "0.3", "150", "1000"],
        ["MT400", "EMISION", "0", "999999999", "EUR", "MX", "ES", "180", "0", "120", "600"],

        # MT410 - Acuse de Recibo de Cobranza
        ["MT410", "RECEPCION", "0", "999999999", "USD", "MX", "US", "50", "0", "30", "200"],
        ["MT410", "RECEPCION", "0", "999999999", "EUR", "MX", "ES", "60", "0", "40", "250"],

        # MT412 - Aviso de Aceptación
        ["MT412", "ACEPTACION", "0", "50000", "USD", "MX", "US", "100", "0", "75", "400"],
        ["MT412", "ACEPTACION", "50000", "999999999", "USD", "MX", "US", "0", "0.2", "100", "800"],

        # MT416 - Aviso de Pago
        ["MT416", "PAGO", "0", "50000", "USD", "MX", "US", "80", "0", "50", "300"],
        ["MT416", "PAGO", "50000", "999999999", "USD", "MX", "US", "0", "0.15", "80", "600"],

        # MT420 - Rastreo de Cobranza
        ["MT420", "RASTREO", "0", "999999999", "USD", "MX", "US", "25", "0", "15", "100"],

        # === CRÉDITOS DOCUMENTARIOS (MT7xx) ===
        # MT700 - Emisión de Carta de Crédito
        ["MT700", "EMISION", "0", "100000", "USD", "MX", "US", "250", "0", "200", "1000"],
        ["MT700", "EMISION", "100000", "500000", "USD", "MX", "US", "0", "0.25", "250", "2000"],
        ["MT700", "EMISION", "500000", "999999999", "USD", "MX", "US", "0", "0.20", "500", "5000"],
        ["MT700", "EMISION", "0", "100000", "EUR", "MX", "ES", "300", "0", "250", "1200"],
        ["MT700", "EMISION", "100000", "999999999", "EUR", "MX", "ES", "0", "0.28", "300", "2500"],

        # MT701 - Emisión de Carta de Crédito Stand-by
        ["MT701", "EMISION", "0", "100000", "USD", "MX", "US", "300", "0", "250", "1200"],
        ["MT701", "EMISION", "100000", "999999999", "USD", "MX", "US", "0", "0.30", "300", "2500"],

        # MT707 - Modificación de Carta de Crédito
        ["MT707", "MODIFICACION", "0", "999999999", "USD", "MX", "US", "150", "0", "100", "800"],
        ["MT707", "MODIFICACION", "0", "999999999", "EUR", "MX", "ES", "180", "0", "120", "900"],

        # MT710 - Aviso de Carta de Crédito de Tercero
        ["MT710", "AVISO", "0", "100000", "USD", "MX", "US", "200", "0", "150", "800"],
        ["MT710", "AVISO", "100000", "999999999", "USD", "MX", "US", "0", "0.20", "200", "1500"],

        # MT711 - Aviso de Carta de Crédito Stand-by de Tercero
        ["MT711", "AVISO", "0", "100000", "USD", "MX", "US", "250", "0", "180", "900"],
        ["MT711", "AVISO", "100000", "999999999", "USD", "MX", "US", "0", "0.22", "250", "1800"],

        # MT720 - Transferencia de Carta de Crédito
        ["MT720", "TRANSFERENCIA", "0", "100000", "USD", "MX", "US", "200", "0", "150", "900"],
        ["MT720", "TRANSFERENCIA", "100000", "999999999", "USD", "MX", "US", "0", "0.18", "200", "1500"],

        # MT730 - Acuse de Recibo
        ["MT730", "RECEPCION", "0", "999999999", "USD", "MX", "US", "50", "0", "30", "200"],
        ["MT730", "RECEPCION", "0", "999999999", "EUR", "MX", "ES", "60", "0", "40", "250"],

        # MT732 - Aviso de Incumplimiento/Rechazo
        ["MT732", "RECHAZO", "0", "999999999", "USD", "MX", "US", "100", "0", "75", "400"],
        ["MT732", "RECHAZO", "0", "999999999", "EUR", "MX", "ES", "120", "0", "90", "480"],

        # MT734 - Aviso de Renuncia
        ["MT734", "RENUNCIA", "0", "999999999", "USD", "MX", "US", "75", "0", "50", "300"],

        # MT740 - Autorización para Reembolsar
        ["MT740", "AUTORIZACION", "0", "100000", "USD", "MX", "US", "120", "0", "80", "500"],
        ["MT740", "AUTORIZACION", "100000", "999999999", "USD", "MX", "US", "0", "0.12", "120", "1000"],

        # MT742 - Reclamación de Reembolso
        ["MT742", "RECLAMACION", "0", "100000", "USD", "MX", "US", "100", "0", "70", "450"],
        ["MT742", "RECLAMACION", "100000", "999999999", "USD", "MX", "US", "0", "0.10", "100", "800"],

        # MT747 - Modificación de Carta de Crédito
        ["MT747", "MODIFICACION", "0", "999999999", "USD", "MX", "US", "150", "0", "100", "800"],
        ["MT747", "MODIFICACION", "0", "999999999", "EUR", "MX", "ES", "180", "0", "120", "900"],

        # MT750 - Aviso de Discrepancia
        ["MT750", "DISCREPANCIA", "0", "999999999", "USD", "MX", "US", "80", "0", "50", "350"],
        ["MT750", "DISCREPANCIA", "0", "999999999", "EUR", "MX", "ES", "100", "0", "60", "400"],

        # MT752 - Autorización para Pagar/Aceptar/Negociar
        ["MT752", "AUTORIZACION", "0", "100000", "USD", "MX", "US", "150", "0", "100", "600"],
        ["MT752", "AUTORIZACION", "100000", "999999999", "USD", "MX", "US", "0", "0.15", "150", "1200"],

        # MT754 - Aviso de Pago/Aceptación
        ["MT754", "PAGO", "0", "100000", "USD", "MX", "US", "120", "0", "80", "500"],
        ["MT754", "PAGO", "100000", "999999999", "USD", "MX", "US", "0", "0.12", "120", "1000"],

        # MT756 - Aviso de Reembolso
        ["MT756", "REEMBOLSO", "0", "100000", "USD", "MX", "US", "100", "0", "70", "450"],
        ["MT756", "REEMBOLSO", "100000", "999999999", "USD", "MX", "US", "0", "0.10", "100", "800"],

        # === GARANTÍAS BANCARIAS (MT7xx) ===
        # MT760 - Garantía Bancaria
        ["MT760", "EMISION", "0", "200000", "USD", "MX", "US", "300", "0", "250", "1500"],
        ["MT760", "EMISION", "200000", "1000000", "USD", "MX", "US", "0", "0.15", "300", "3000"],
        ["MT760", "EMISION", "1000000", "999999999", "USD", "MX", "US", "0", "0.12", "500", "5000"],
        ["MT760", "EMISION", "0", "200000", "EUR", "MX", "ES", "350", "0", "280", "1600"],
        ["MT760", "EMISION", "200000", "999999999", "EUR", "MX", "ES", "0", "0.18", "350", "3500"],

        # MT767 - Modificación de Garantía
        ["MT767", "MODIFICACION", "0", "999999999", "USD", "MX", "US", "200", "0", "150", "1000"],
        ["MT767", "MODIFICACION", "0", "999999999", "EUR", "MX", "ES", "250", "0", "180", "1200"],

        # === MENSAJES ADICIONALES COMERCIO EXTERIOR ===
        # MT799 - Mensaje de Texto Libre entre Bancos
        ["MT799", "CONSULTA", "0", "999999999", "USD", "MX", "US", "30", "0", "20", "100"],
        ["MT799", "CONSULTA", "0", "999999999", "EUR", "MX", "ES", "35", "0", "25", "120"],
    ]

    for row_idx, row_data in enumerate(data_examples, start=10):
        for col_idx, value in enumerate(row_data, start=1):
            # Convert integer strings to float for commission columns (8-11)
            if col_idx >= 8 and value.replace('.', '', 1).isdigit():
                value = float(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Ajustar anchos de columna
    column_widths = [15, 18, 15, 15, 12, 15, 15, 15, 18, 17, 17]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Ajustar altura de fila 9 (constraints)
    ws.row_dimensions[9].height = 30

    # Guardar archivo
    output_path = "src/main/resources/comisiones-swift-config.xlsx"
    wb.save(output_path)
    print(f"✓ Archivo creado exitosamente: {output_path}")
    print(f"✓ Contiene {len(data_examples)} reglas de ejemplo")
    print(f"✓ Formato compatible con Drools Decision Table")

    return output_path

if __name__ == "__main__":
    try:
        crear_archivo_comisiones()
    except ImportError:
        print("ERROR: openpyxl no está instalado")
        print("Instalar con: pip install openpyxl")
    except Exception as e:
        print(f"ERROR: {e}")
