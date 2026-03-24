#!/usr/bin/env python3
import openpyxl

# Cargar el archivo Excel
wb = openpyxl.load_workbook('/Users/admin/Documents/capacitacion/claude/claudecode/globalcmx/backend/src/main/resources/comisiones-swift-config.xlsx')

print('Hojas disponibles:', wb.sheetnames)
ws = wb.active
print('Hoja activa:', ws.title)
print('\nPrimeras 15 filas:')
print('-' * 100)

for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
    print(f'Fila {i}: {row}')

print('\nTotal de filas con datos:', ws.max_row)
print('Total de columnas con datos:', ws.max_column)
